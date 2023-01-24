import openpyxl

wb = openpyxl.Workbook()

print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet')

sheet['A1'] = 42
sheet['A2'] = 'Hello'

sheet2 = wb.create_sheet()
sheet2.title = 'My New Sheet Name'
print(wb.get_sheet_names())

wb.create_sheet(index=0, title='My Other Sheet')

wb.save('example2.xlsx')
